
import os
import datetime
import shutil
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side

#  BACK-END


# Caminhos fixos
CAMINHO_PLANILHAS = "C:/Comparador de Planilhas/Viagens do dia/"
CAMINHO_BACKUP = "C:/Unificador de Planilhas/Backup das viagens/"
CAMINHO_SAIDA = "C:/Unificador de Planilhas/Viagens do dia - Geral/"

# Lista de arquivos selecionados
arquivos_selecionados = []

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side

def copiar_celula(orig_cell, dest_cell):
    dest_cell.value = orig_cell.value
    dest_cell.value = orig_cell.value

def selecionar_planilhas():
    global arquivos_selecionados
    if not os.path.exists(CAMINHO_PLANILHAS):
        messagebox.showerror("Erro", f"Caminho não encontrado:\n{CAMINHO_PLANILHAS}")
        return

    arquivos_selecionados.clear()
    arquivos_selecionados += [
        os.path.join(CAMINHO_PLANILHAS, f)
        for f in os.listdir(CAMINHO_PLANILHAS)
        if f.lower().endswith(".xlsx")
    ]

    if arquivos_selecionados:
        status_var.set(f"✔ {len(arquivos_selecionados)} planilha(s) carregadas.")
    else:
        status_var.set("Nenhuma planilha .xlsx encontrada na pasta.")

def abrir_pasta_saida():
    if not os.path.exists(CAMINHO_SAIDA):
        os.makedirs(CAMINHO_SAIDA)
    os.startfile(CAMINHO_SAIDA)

def unificar_planilhas_formatadas():
    if not arquivos_selecionados:
        messagebox.showwarning("Aviso", "Nenhuma planilha carregada.")
        return

    data_hoje = datetime.datetime.now().strftime("%d-%m-%Y")

    if not os.path.exists(CAMINHO_SAIDA):
        os.makedirs(CAMINHO_SAIDA)

    base_nome = f"Planilha Geral {data_hoje}"
    contador = 1
    caminho_saida = os.path.join(CAMINHO_SAIDA, f"{base_nome}.{contador}.xlsx")
    while os.path.exists(caminho_saida):
        contador += 1
        caminho_saida = os.path.join(CAMINHO_SAIDA, f"{base_nome}.{contador}.xlsx")

    status_var.set("Processando...")

    dados_registro = set()
    dados_linhas = []
    ws_headers = None

    for arquivo in arquivos_selecionados:
        wb = load_workbook(arquivo)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        if not ws_headers:
            ws_headers = [cell for cell in ws[1]]

        turno_idx = next((i for i, v in enumerate(header) if "turno" in str(v).lower()), None)
        itinerario_idx = next((i for i, v in enumerate(header) if "itin" in str(v).lower()), None)
        registro_idx = next((i for i, v in enumerate(header) if "registro" in str(v).lower()), None)

        for row in ws.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            if (
                not values[registro_idx] or
                not values[itinerario_idx] or
                str(values[itinerario_idx]).lower() == "itinerário" or
                str(values[turno_idx]).lower().startswith("turno") or
                not str(values[itinerario_idx]).replace('.', '', 1).replace(',', '', 1).isdigit()
            ):
                continue

            registro = values[registro_idx]
            itinerario = values[itinerario_idx]
            turno = values[turno_idx]

            if registro in dados_registro:
                continue
            dados_registro.add(registro)

            dados_linhas.append({
                "turno": turno,
                "itinerario": itinerario,
                "registro": registro,
                "row": row
            })

    dados_ordenados = sorted(
        dados_linhas,
        key=lambda x: (
            int(x["turno"]) if str(x["turno"]).isdigit() else 0,
            float(str(x["itinerario"]).replace(',', '.')) if x["itinerario"] else 0
        )
    )

    wb_final = Workbook()
    ws_final = wb_final.active
    ws_final.title = data_hoje
    linha_final = 1
    ultimo_turno = None
    itinerarios_adicionados = set()
    primeiro_bloco = True

    for item in dados_ordenados:
        turno = item["turno"]
        itinerario = item["itinerario"]
        row = item["row"]

        if turno != ultimo_turno:
            if not primeiro_bloco:
                ws_final.append([""] * len(row))
                linha_final += 1
                for col in range(1, len(row)+1):
                    ws_final.cell(row=linha_final, column=col).fill = PatternFill(start_color="5591f9", end_color="5591f9", fill_type="solid")
                linha_final += 1
                ws_final.append([""] * len(row))
                linha_final += 1

            ws_final.append([""] * len(row))
            linha_turno = ws_final.max_row
            cell_turno = ws_final.cell(row=linha_turno, column=1)
            cell_turno.value = f"Turno {turno}"
            cell_turno.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            cell_turno.font = Font(bold=True)
            cell_turno.alignment = Alignment(horizontal="center")
            for col in range(1, len(row)+1):
                cell = ws_final.cell(row=linha_turno, column=col)
                cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            linha_final = linha_turno + 1
            ultimo_turno = turno
            itinerarios_adicionados = set()
            primeiro_bloco = False

        if itinerario not in itinerarios_adicionados:
            ws_final.append([""] * len(row))
            linha_final += 1
            for j, cell in enumerate(ws_headers):
                cell_header = ws_final.cell(row=linha_final, column=j+1)
                cell_header.value = cell.value
                cell_header.font = Font(bold=True)
                cell_header.border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin")
                )
                if j < 4:
                    cell_header.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # amarelo
                else:
                    cell_header.fill = PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid")  # azul claro
            linha_final += 1
            itinerarios_adicionados.add(itinerario)

        for j, cell in enumerate(row):
            dest = ws_final.cell(row=linha_final, column=j+1)
            copiar_celula(cell, dest)
            dest.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
        linha_final += 1

    wb_final.save(caminho_saida)

    # Após salvar a planilha, mover os arquivos de entrada para o backup
    if not os.path.exists(CAMINHO_BACKUP):
        os.makedirs(CAMINHO_BACKUP)
    for caminho in arquivos_selecionados:
        nome_arquivo = os.path.basename(caminho)
        destino_arquivo = os.path.join(CAMINHO_BACKUP, nome_arquivo)
        count = 1
        while os.path.exists(destino_arquivo):
            nome_base, ext = os.path.splitext(nome_arquivo)
            destino_arquivo = os.path.join(CAMINHO_BACKUP, f"{nome_base}_{count}{ext}")
            count += 1
        shutil.move(caminho, destino_arquivo)

    status_var.set(f"✔ Planilha gerada: {os.path.basename(caminho_saida)}")
    messagebox.showinfo("Sucesso", f"Arquivo salvo:\n{caminho_saida}")





# FRONT-END
class App:
    COR_PRINCIPAL = "#eaf4fb"
    COR_BOTAO = "#007acc"
    COR_BOTAO_HOVER = "#005f99"
    COR_TEXTO = "#003b5c"
    COR_STATUS = "#1c4966"

    def __init__(self, root):
        self.root = root
        self.root.title("Unificador de Planilhas")
        self.root.geometry("500x350")
        self.root.configure(bg=self.COR_PRINCIPAL)

        self.status_var = tk.StringVar()
        self.status_var.set("Nenhuma planilha carregada ainda.")
        global status_var
        status_var = self.status_var  # permite usar no backend

        self.criar_interface()

    def criar_interface(self):
        tk.Label(
            self.root,
            text="Unificador de Planilhas Excel",
            bg=self.COR_PRINCIPAL,
            fg=self.COR_TEXTO,
            font=("Segoe UI", 16, "bold")
        ).pack(pady=20)

        self.criar_botao("Carregar planilhas da pasta", selecionar_planilhas).pack(pady=5)
        self.criar_botao("Unificar e Salvar", unificar_planilhas_formatadas).pack(pady=5)
        self.criar_botao("Abrir pasta de saída", abrir_pasta_saida).pack(pady=5)
        self.criar_botao("Sair", self.root.destroy).pack(pady=5)

        tk.Label(
            self.root,
            textvariable=self.status_var,
            bg=self.COR_PRINCIPAL,
            fg=self.COR_STATUS,
            font=("Segoe UI", 10, "italic"),
            wraplength=480
        ).pack(pady=15)

    def criar_botao(self, texto, comando):
        btn = tk.Button(
            self.root,
            text=texto,
            command=comando,
            bg=self.COR_BOTAO,
            fg="white",
            activebackground=self.COR_BOTAO_HOVER,
            activeforeground="white",
            font=("Segoe UI", 10, "bold"),
            width=30,
            height=2,
            bd=0,
            relief="flat",
            cursor="hand2"
        )
        btn.bind("<Enter>", lambda e: btn.config(bg=self.COR_BOTAO_HOVER))
        btn.bind("<Leave>", lambda e: btn.config(bg=self.COR_BOTAO))
        return btn


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
